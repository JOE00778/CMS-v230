"""JD/BM 商品登録模板 schema + NST 主档行 → JD/BM 行映射

下载模板（Boss 2026-06-23 更新·严格按格式·含合并单元格·尤重 row1/row2）：
- JD（京东）：Import-BasicGoods-SinglePage-Chinese.xlsx · sheet「商品信息」(75 列)
- BM（斑马）：Product导入模板.xlsx · sheet「数据」(46 列)
两者 row1=分区标题（合并单元格）/ row2=列头 / row3+=数据。生成与模板逐格一致（脚本校验）。

映射策略（Boss 2026-05-29 拍板）：
- JD *货主编码：默认 "KH20000009340"（旧 HTML 商品登録ツール jdA 默认值）
- BM SPU：使用 NST JANコード（13 位纯数字、唯一）
- BM ERP 类目：留空，用户后填
- 英文名称（JD/BM 共）：留空，用户用别的工具翻

衍生规则：从 NST 主档行（dict, key=NST 模板列名）单行生成 JD/BM 行。
图片 URL 由调用方传入（来自 nst.item_image_cache 的查询结果）。
"""
from __future__ import annotations

import datetime
import io
from typing import Iterable

from shared.jp_translit import to_english_title

# ───────────────────────── JD 商品信息 schema（新 单页基础商品导入·2026-06-23） ─────────────────────────
# 源模板：Import-BasicGoods-SinglePage-Chinese.xlsx · sheet「商品信息」(75 列)
# row1=分区标题（合并单元格·仅区首列有值）/ row2=列表头 / row3+=数据
# Boss 样本(35 行)实填的 8 列：客户SKU/商品名称/件型=1/自带原包=1/商品条码/销售渠道/平台编码/平台标题
JD_SHEET_NAME = "商品信息"

JD_GROUP = [
    '商品基本信息',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '商家商品件型',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '商品仓储物流信息',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '商品条码',
    '销售渠道商品信息',
    '',
    '',
    '',
    '',
    '',
    '多级包装（若无此业务场景，请不要填写本sheet任何内容）',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
]

JD_HEADER = [
    '货主ID\n(客户只有单货主则不必填）',
    '*客户SKU',
    '*商品名称',
    '*件型\n（0-大件 \n1-中小件 \n2-小件 \n3-中件 ）',
    '三级品类编码',
    '规格型号',
    '颜色',
    '款号',
    '商品报关价格',
    '商品报关币种',
    '海关关税代码',
    '品牌名称（中文）',
    '品牌名称（英文）',
    'ISO标准国家代码',
    '原产国/地区\n（原产国信息数据请见界面）',
    '原产地区\n（原产地区信息数据请见界面）',
    '生产企业名称',
    '生产企业地址',
    '供应商',
    '计量单位\n（单位信息数据请见界面）',
    '自带原包\n（0-否；1-是）',
    '毛重',
    '净重',
    '带电组件净重',
    '重量单位\n（kg, g, lb）',
    '长',
    '宽',
    '高',
    '长度单位\n（cm, mm, in）',
    '电池容量\n（单位默认WH）',
    '打包方式',
    '备注',
    '是否序列号管理\n（0-否；1-是）',
    '是否批次管理\n（0-否；1-是）',
    '是否入库批次管理\n（0-否；1-是）',
    '是否包装批号管理\n（0-否；1-是）',
    '是否保质期管理\n（0-否；1-生产日期管理；2-到期日期管理）',
    '保质期天数\n（大于0的整数）',
    '允许入库天数\n（过了生产日期多久之内 允许入库）',
    '临期天数\n（距离到期多少天， 商品变为临期状态）',
    '预警天数',
    '安全库存数',
    '入库日期管理（0-否，1-日，2-半月，3-月，4-季度，5-半年，6-年； 默认值0）',
    '是否轻量化序列号管理\n（0-否；1-是）',
    '序列号规则类型\n（0-正向规则(默认)；1-反向规则）\n序列号规则，此处只支持一个商品导入一条规则，如果要维护多条规则，请在FOP商品页面维护，或者序列号规则页面维护或批量导入；',
    '序列号总长度',
    '前缀长度',
    '前缀内容',
    '序列号格式\n0-无限制；1-仅字母；2-仅数字；3-字母和数字；4-字母和特殊字符；5-数字和特殊字符',
    '是否危险品管理（0-否；1-是）',
    'UN CODE\n(1-UN3480; 2-UN3481; 3-UN1266; 4-UN3090; 5-UN3091; 6-UN3171; 7-UN1123; 8-UN1169; 9-UN1170; 10-UN1230; 11-UN1263; 12-UN1770; 13-UN1950; 14-UN1987; 15-UN1993; 16-UN3082; 17-UN3175)',
    '是否耗材管理\n（1:是 0:否）',
    '是否易碎品\n（1:是 0:否）',
    '商品尺寸',
    '是否异形品\n（1:是 0:否）',
    '*商品条码',
    '*销售渠道编码',
    '*平台商品编码',
    '*平台商品标题',
    '平台客户SKU',
    '平台规格型号',
    '平台商品链接',
    '*包装代码管理\n（1-一级、2-二级、3-三级）\n代表商品在按件管理之上，一共还需要几级包装的管理，则在本单元格填写数字。例如商品按箱+盒+件的管理，则属于一共2级的包装，需要填写2行记录，且本单元格都填2，并分别将每级的信息填写至2行记录中对应的后续其他各列）',
    '*包装等级\n代表本行记录填写的信息属于第几级包装\n（例如，如果件之上一共有两级包装。则本单元格填1，代表件之上的第1级，属于最小包装等级，本行记录中后续各多级包装信息列都属于第1级的内容；本单元格填2，代表第2级包装，本行记录中后续各多级包装信息列都属于第1级的内容）',
    '*包装名称\n（1-箱，2-盒，3-包，4-袋）\n代表本行记录中填写的包装等级，客户命名的单位名称',
    '*次级包装数量\n（填写大于1的正整数值，代表本行记录填写的包装级别包含的下一级包装的数量，例如本行记录属于第2级包装，则本单元格填写的整数值，代第2级包含的第1级包装数量。若本行记录的信息属于第1级包装，此时本单元格填写的值代表包含的商品件数）',
    '客户包装条码',
    '长',
    '宽',
    '高',
    '长度单位\n(1-cm 2-mm 3-in)\n若填写了I J、K三列，则必须填长度单位',
    '重量',
    '重量单位\n(1-kg、2-g、3-lb)\n若填写了M列，则必须填重量单位',
    '质检图片',
    '质检说明',
]

assert len(JD_GROUP) == 75, f"JD_GROUP={len(JD_GROUP)}"
assert len(JD_HEADER) == 75, f"JD_HEADER={len(JD_HEADER)}"

DEFAULT_JD_CUSTOMER_CODE = "KH20000009340"   # 货主ID（新模板可不填·单货主）· 默认输入留空
DEFAULT_JD_SALES_CHANNEL = "sc-三金商事株式会社"
DEFAULT_JD_PLATFORM_CODE = "Lazada/Shopee/coupang"

# row1 分区标题の合并单元格（テンプレと完全一致·Boss 2026-06-23「严格按格式·含合并」）
JD_MERGES = ["A1:U1", "V1:AE1", "AF1:BC1", "BE1:BJ1", "BK1:BU1"]

# ───────────────────────── BM「Product导入模板」schema（新·2026-06-23） ─────────────────────────
# 源模板：Product导入模板.xlsx · sheet「数据」(46 列)
# row1=分区標題（合并单元格）/ row2=列頭（旧 BM_HEADER と同一·映射不変）/ row3+=データ
BM_SHEET_NAME = "数据"

BM_GROUP = (
    ["SPU(相同信息可以填写一样的)"] + [""] * 11    # 1-12  A1:L1
    + ["SKU"] + [""] * 16                          # 13-29 M1:AC1
    + ["普通报关信息"] + [""] * 5                   # 30-35 AD1:AI1
    + ["报关特殊属性"] + [""] * 7                   # 36-43 AJ1:AQ1
    + ["图片信息"]                                 # 44    AR1（単独）
    + ["质检制作要求", ""]                          # 45-46 AS1:AT1
)
assert len(BM_GROUP) == 46, len(BM_GROUP)

BM_MERGES = ["A1:L1", "M1:AC1", "AD1:AI1", "AJ1:AQ1", "AS1:AT1"]

BM_HEADER = [
    "SPU", "产品标题", "ERP类目", "来源URL", "来源备注", "默认供应商名称",
    "富文本描述", "纯文本描述", "短描述", "SEO标题", "SEO关键字", "SEO描述",
    "SKU", "图片URL",
    "规格1名称", "规格1值", "规格2名称", "规格2值", "规格3名称", "规格3值",
    "规格4名称", "规格4值", "规格5名称", "规格5值",
    "成本价(￥)", "重量(g)", "长(cm)", "宽(cm)", "高(cm)",
    "中文名称", "英文名称",
    "材质", "申报价值(USD)", "报关重量(g)", "海关编码",
    "带电(非内置)", "带电(内置)", "带电(纯电池)", "带磁",
    "液体", "粉末", "刀具", "危险品",
    "产品图片(URL)",
    "项目名", "标准描述",
]


# ───────────────────────── NST → JD/BM 行映射 ─────────────────────────

def _g(nst_row: dict, *keys, default=""):
    """安全取 NST 行字段（多个候选 key 选第一个非空）。"""
    for k in keys:
        v = nst_row.get(k)
        if v is not None and str(v).strip() != "" and str(v).strip().lower() != "nan":
            return v
    return default


def _num_or_blank(v) -> str:
    if v in (None, ""): return ""
    s = str(v).strip()
    if s == "" or s.lower() == "nan": return ""
    return s


def nst_to_jd_row(
    nst_row: dict,
    *,
    image_url: str = "",
    jd_customer_code: str = DEFAULT_JD_CUSTOMER_CODE,
    sales_channel: str = DEFAULT_JD_SALES_CHANNEL,
    platform_code: str = DEFAULT_JD_PLATFORM_CODE,
) -> list:
    """NST 行 → 新「商品信息」行（list, len=75, 顺序 = JD_HEADER）。

    Boss 样本(Import-BasicGoods-SinglePage-Chinese)实填 8 列：
      col1 货主ID=jd_customer_code（空=单货主不必填）· col2 客户SKU=JAN · col3 商品名称=日文名
      · col4 件型=1（中小件）· col21 自带原包=1 · col56 商品条码=JAN · col57 销售渠道 · col58 平台编码
      · col59 平台商品标题=英文（留空·用户后填，我们无英文源）
    毛重/尺寸/规格型号 等留空（新模板这些不要了）。image_url 本模板不用。
    """
    jan = _g(nst_row, "JANコード")
    name_ja = _g(nst_row, "アイテム名")
    row = [""] * len(JD_HEADER)
    row[0]  = jd_customer_code   # col1  货主ID（空=不填）
    row[1]  = jan                # col2  *客户SKU
    row[2]  = name_ja            # col3  *商品名称（日文·用户可改）
    row[3]  = "1"                # col4  *件型（1=中小件）
    row[20] = "1"                # col21 自带原包（1=是）
    row[55] = jan                # col56 *商品条码
    row[56] = sales_channel      # col57 *销售渠道编码
    row[57] = platform_code      # col58 *平台商品编码
    row[58] = _g(nst_row, "英文标题") or to_english_title(name_ja)  # col59 平台商品标题(英文·离线转写草稿)
    return row


def nst_to_bm_row(
    nst_row: dict,
    *,
    image_url: str = "",
) -> list:
    """NST 行 → BM 行（list, len=46, 顺序 = BM_HEADER）。"""
    jan = _g(nst_row, "JANコード")
    name_ja = _g(nst_row, "アイテム名")
    cost = _g(nst_row, "商品原価")
    weight_g = _g(nst_row, "商品重量(g)", "パッケージ重量(g)")
    length = _g(nst_row, "商品奥行(cm)", "パッケージ奥行(cm)")
    width = _g(nst_row, "商品幅(cm)", "パッケージ幅(cm)")
    height = _g(nst_row, "商品高さ(cm)", "パッケージ高さ(cm)")

    row = [""] * len(BM_HEADER)
    row[0]  = jan                         # SPU = JAN (Boss 拍板)
    row[1]  = name_ja                     # 产品标题
    row[2]  = ""                          # ERP类目（留空，Boss 拍板）
    row[12] = jan                         # SKU
    row[13] = image_url                   # 图片URL
    row[14] = "1"                         # 规格1名称（sample 模板值）
    row[15] = jan                         # 规格1值（sample 模板值）
    row[24] = _num_or_blank(cost)         # 成本价(￥)
    row[25] = _num_or_blank(weight_g)     # 重量(g)
    row[26] = _num_or_blank(length)       # 长(cm)
    row[27] = _num_or_blank(width)        # 宽(cm)
    row[28] = _num_or_blank(height)       # 高(cm)
    row[29] = name_ja                     # 中文名称（暂用日文）
    row[30] = _g(nst_row, "英文标题") or to_english_title(name_ja)  # 英文名称(离线转写草稿)
    row[43] = image_url                   # 产品图片(URL)
    return row


# ───────────────────────── xlsx 生成（openpyxl） ─────────────────────────

def _build_xlsx(sheet_name: str, group: list[str], header: list[str],
                data_rows: list[list], merges: list[str] | None = None) -> bytes:
    """生成 xlsx bytes · row1=分区标题（含合并单元格）/ row2=列头 / row3+=数据。

    merges: 'A1:U1' 等の合并范围リスト（テンプレ row1 分区と一致させる）。
    合并時は非アンカーセルを None にしてから merge（openpyxl 警告回避）。
    """
    from openpyxl import Workbook
    from openpyxl.utils import range_boundaries
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(group)
    ws.append(header)
    for r in data_rows:
        ws.append(r)
    for rng in (merges or []):
        c1, r1, c2, r2 = range_boundaries(rng)
        for col in range(c1, c2 + 1):           # アンカー以外を空に（merge 警告回避）
            for row in range(r1, r2 + 1):
                if not (col == c1 and row == r1):
                    ws.cell(row=row, column=col).value = None
        ws.merge_cells(rng)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_jd_xlsx(
    nst_rows: Iterable[dict],
    *,
    image_url_map: dict[str, str] | None = None,
    jd_customer_code: str = DEFAULT_JD_CUSTOMER_CODE,
    sales_channel: str = DEFAULT_JD_SALES_CHANNEL,
    platform_code: str = DEFAULT_JD_PLATFORM_CODE,
) -> bytes:
    image_url_map = image_url_map or {}
    data = []
    for n in nst_rows:
        jan = str(_g(n, "JANコード") or "").strip()
        data.append(nst_to_jd_row(
            n,
            image_url=image_url_map.get(jan, ""),
            jd_customer_code=jd_customer_code,
            sales_channel=sales_channel,
            platform_code=platform_code,
        ))
    return _build_xlsx(JD_SHEET_NAME, JD_GROUP, JD_HEADER, data, merges=JD_MERGES)


def build_bm_xlsx(
    nst_rows: Iterable[dict],
    *,
    image_url_map: dict[str, str] | None = None,
) -> bytes:
    image_url_map = image_url_map or {}
    data = []
    for n in nst_rows:
        jan = str(_g(n, "JANコード") or "").strip()
        data.append(nst_to_bm_row(n, image_url=image_url_map.get(jan, "")))
    return _build_xlsx(BM_SHEET_NAME, BM_GROUP, BM_HEADER, data, merges=BM_MERGES)


def dated_filename_jd() -> str:
    return f"JD_{datetime.date.today():%Y%m%d}.xlsx"


def dated_filename_bm() -> str:
    return f"BM_{datetime.date.today():%Y%m%d}.xlsx"
