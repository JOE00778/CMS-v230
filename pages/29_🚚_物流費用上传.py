"""模块 #29 物流費用上传 · JD 請求書 + BM(包裹号→店舗) を上传 → 配賦集計。

Boss 指示:
- 費用請求は上传のみで録入（NST 拉取不可）。
- 订单→店舗 マッピング(BM) も NST 拉取不可 → Boss が平台/JD WMS から導出 upload。
- 包裹号で紐付かない / 店舗が部署未分類 → 【不明】に集計（後で確認）。
- 列は表頭名で解決（JD 請求書/BM は列順不定）。費用は不含税。
- 2026-08-29（Boss 拍板の最終形態）: 請求書 upload 時に**斑马 API から包裹→店舗を
  自動補齊**（BM 手動導出を代替 · tab1 の BM upload は後備として残す）。
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta, timezone

import openpyxl
import pandas as pd
import streamlit as st

from shared import banma_client, nst_suiteql
from shared.db import get_connection
from shared.i18n import lang_selector, t

st.set_page_config(page_title=t("物流费用上传"), page_icon="🚚", layout="wide")
from shared.auth import require_password, require_extra_password
from shared.theme import inject_theme
require_password()
require_extra_password("sys", "SYS_SETTINGS_PW", default="1001")  # 系统设置二级密码
inject_theme()
lang_selector()
conn = get_connection()

st.title(t("🚚 物流费用上传"))
st.caption(t(
    "上传 JD 请求书（费用明细）与 BM（包裹号→店铺）→ 自动配賦汇总。"
    "费用为不含税 · 无法关联的计入【不明】。结果在「🚚 物流费用分析」查看。"
))

INV_SHEETS = {
    "OB-Pick&Pack":   ("pickpack", ("客户出库单号",)),
    "Last mile":      ("lastmile", ("客户运单编码",)),
    "Packing Charge": ("packing",  ("客户出库单编码",)),
}


def find_col(hdr, *kw, exclude=None):
    for i, h in enumerate(hdr):
        if not h:
            continue
        s = str(h)
        if any(k in s for k in kw) and (exclude is None or exclude not in s):
            return i
    return None


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
        if m:
            return date(int(m[1]), int(m[2]), int(m[3]))
    return None


def _cell(r, c):
    v = r[c] if (c is not None and c < len(r)) else None
    return v


def parse_invoice(data: bytes, ym: str):
    # 普通模式（請求書/导出 xlsx は dimension 欠落で read_only 不可の場合あり）
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    rows, per = [], {}
    for sheet, (ct, joinkw) in INV_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        it = wb[sheet].iter_rows(values_only=True)
        next(it, None)
        hdr = list(next(it, []))
        cj = find_col(hdr, *joinkw)
        ce = find_col(hdr, "不含税")
        ci = find_col(hdr, "含税金额", exclude="不含税")
        cd = find_col(hdr, "费用发生时间", "費用発生")
        cs = find_col(hdr, "商品编号", "SKU") if ct == "pickpack" else None
        cmc = find_col(hdr, "耗材编码") if ct == "packing" else None
        cmq = find_col(hdr, "耗材数量") if ct == "packing" else None
        n = 0
        for r in it:
            jk = _cell(r, cj)
            amt = _num(_cell(r, ce))
            if jk is None and amt is None:
                continue
            rows.append((
                ym, ct,
                str(jk).strip() if jk is not None else None,
                amt, _num(_cell(r, ci)),
                (str(_cell(r, cmc)).strip() if _cell(r, cmc) else None),
                _num(_cell(r, cmq)),
                (str(_cell(r, cs)).strip() if _cell(r, cs) else None),
                _to_date(_cell(r, cd)),
            ))
            n += 1
        per[ct] = n
    wb.close()
    return rows, per


def parse_bm(data: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = None
    for s in wb.worksheets:
        h = list(next(s.iter_rows(values_only=True, max_row=1), []))
        if find_col(h, "包裹号", "包裹號") is not None:
            ws = s
            break
    if ws is None:
        wb.close()
        return []
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    cp = find_col(hdr, "包裹号", "包裹號")
    co = find_col(hdr, "订单号", "訂単")
    cw = find_col(hdr, "物流单号", "物流單号")
    cpl = find_col(hdr, "平台")
    csh = find_col(hdr, "店铺", "店舗")
    cdt = find_col(hdr, "发货时间", "發貨", "出荷")
    seen = {}
    for r in it:
        pk = _cell(r, cp)
        if not pk:
            continue
        key = str(pk).strip()
        seen[key] = (
            key,
            str(_cell(r, co)).strip() if _cell(r, co) else None,
            str(_cell(r, cw)).strip() if _cell(r, cw) else None,
            str(_cell(r, cpl)).strip() if _cell(r, cpl) else None,
            str(_cell(r, csh)).strip() if _cell(r, csh) else None,
            _to_date(_cell(r, cdt)),
        )
    wb.close()
    return list(seen.values())


def parse_billing(data: bytes, ym: str):
    """Billing sheet（JD 公式の費用全構成）→ (ym, seq, item_name, ex, in)。Total 行で停止。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    if "Billing" not in wb.sheetnames:
        wb.close()
        return []
    rows = list(wb["Billing"].iter_rows(values_only=True))
    hdr_idx = next((i for i, r in enumerate(rows)
                    if any(str(c).strip() == "Item" for c in r if c)
                    and any("Price before Tax" in str(c) for c in r if c)), None)
    if hdr_idx is None:
        wb.close()
        return []
    hdr = [(str(x).strip() if x else "") for x in rows[hdr_idx]]
    ci_cust = find_col(hdr, "Customer")
    ci_item = find_col(hdr, "Item")
    ci_ex = find_col(hdr, "Price before Tax", "before Tax")
    ci_in = find_col(hdr, "Tax inclusive")
    out, seq = [], 0
    for r in rows[hdr_idx + 1:]:
        cust = str(_cell(r, ci_cust)).strip() if _cell(r, ci_cust) else ""
        item = str(_cell(r, ci_item)).strip() if _cell(r, ci_item) else ""
        if cust.lower() == "total" or item.lower() == "total":
            break
        if not item:
            continue
        ex, inc = _num(_cell(r, ci_ex)), _num(_cell(r, ci_in))
        if ex is None and inc is None:
            continue
        seq += 1
        out.append((ym, seq, item.replace(chr(10), " ").replace(chr(13), " "), ex, inc))
    wb.close()
    return out


def run_recompute(ym=None):
    # 紐付は parcel_no（JD 19桁）優先 → order_id 後備（2026-07-14: Coupang は 2026-05 以降
    # 請求書の出庫単号が Coupang 注文番号(13/14桁)になり parcel_no と不一致 → 全部【不明】化）。
    # order_id は全庫で店舗衝突ゼロ確認済 · サブクエリで order_id 一意化し二重計上を防ぐ。
    conn.execute(
        "DELETE FROM logistics.cost_monthly WHERE (%s IS NULL OR year_month=%s)",
        (ym, ym),
    )
    conn.execute(
        """
        INSERT INTO logistics.cost_monthly
            (year_month, dept, shop, cost_type, amount, qty, source, computed_at)
        SELECT r.year_month,
               COALESCE(d.dept, '不明'),
               COALESCE(COALESCE(mp.shop, mo.shop), '(未マッチ包裹)'),
               r.cost_type,
               SUM(r.amount_ex_tax), COUNT(*), 'recompute', now()
        FROM logistics.cost_invoice_raw r
        LEFT JOIN logistics.order_shop_map mp ON r.join_key = mp.parcel_no
        LEFT JOIN (SELECT order_id, MIN(shop) AS shop
                   FROM logistics.order_shop_map
                   WHERE order_id IS NOT NULL AND shop IS NOT NULL
                   GROUP BY order_id) mo ON r.join_key = mo.order_id
        LEFT JOIN logistics.shop_dept_map  d ON COALESCE(mp.shop, mo.shop) = d.shop
        WHERE (%s IS NULL OR r.year_month = %s)
        GROUP BY r.year_month, COALESCE(d.dept, '不明'),
                 COALESCE(COALESCE(mp.shop, mo.shop), '(未マッチ包裹)'), r.cost_type
        """,
        (ym, ym),
    )
    conn.commit()


def show_match_feedback():
    # 紐付判定は run_recompute と同口径（parcel_no 優先 + order_id 後備）
    rs = conn.execute(
        """SELECT r.year_month AS ym,
                  round(100.0*count(*) FILTER (WHERE mp.parcel_no IS NOT NULL OR mo.order_id IS NOT NULL)/count(*),1) AS match_pct,
                  round(coalesce(sum(r.amount_ex_tax) FILTER (WHERE mp.parcel_no IS NULL AND mo.order_id IS NULL),0))::bigint AS unknown_amt,
                  round(100.0*coalesce(sum(r.amount_ex_tax) FILTER (WHERE mp.parcel_no IS NULL AND mo.order_id IS NULL),0)/nullif(sum(r.amount_ex_tax),0),1) AS unknown_pct
           FROM logistics.cost_invoice_raw r
           LEFT JOIN logistics.order_shop_map mp ON r.join_key = mp.parcel_no
           LEFT JOIN (SELECT DISTINCT order_id FROM logistics.order_shop_map
                      WHERE order_id IS NOT NULL) mo ON r.join_key = mo.order_id
           GROUP BY r.year_month ORDER BY r.year_month"""
    ).fetchall()
    df = pd.DataFrame([dict(x) for x in rs])
    if df.empty:
        return
    df = df.rename(columns={"ym": t("月"), "match_pct": t("件数命中%"),
                            "unknown_amt": t("不明金额(¥)"), "unknown_pct": t("不明率%")})
    st.markdown(t("###### 📊 各月 关联情况（不明率高的月 = 需要补全同期·全平台 BM）"))
    st.dataframe(df, hide_index=True, use_container_width=True,
                 column_config={t("不明金额(¥)"): st.column_config.NumberColumn(format="¥%,.0f")})


def detect_kind(data: bytes, name: str):
    """ファイル種別と対象月を自動判定（請求書=費用sheet有/月はファイル名、BM=包裹号列有）。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    if sheets & {"OB-Pick&Pack", "Last mile", "Packing Charge", "Billing"}:
        m = re.search(r"(20\d{2})\D?(\d{2})", name)
        return "invoice", (f"{m[1]}-{m[2]}" if m else None)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)  # 包裹导出は read_only 不可
    is_bm = any(find_col(list(next(s.iter_rows(values_only=True, max_row=1), [])), "包裹号", "包裹號") is not None
                for s in wb.worksheets)
    wb.close()
    return ("bm", None) if is_bm else ("unknown", None)


tab1, tab2, tab3 = st.tabs([
    t("📤 批量上传（请求书 + BM）"),
    t("店铺→部署 分类"),
    t("🇰🇷 Coupang 费用"),
])

# ============================================================
# tab1 · 一括アップロード（請求書 + BM 自動判定）
# ============================================================
with tab1:
    st.markdown(t("##### 请求书 + BM 一起上传（可多选·自动判定类别与月份）"))
    st.caption(t("请求书: 录入 OB-Pick&Pack/Last mile/Packing/Billing（不含税）· 月份取自文件名的 YYYYMM。"
                 "BM: 包裹号→店铺。两者一起选择批量处理。"))
    st.warning(t(
        "⚠️ BM 需按**对象月同期 × 全平台**导出（国内: 楽天/Amazon/Yahoo/Temu/TikTok ＋ 海外: Shopee/Lazada）。"
        "若仅海外则国内运送费(Last mile)无法关联，【不明】激增。"
    ))
    ups = st.file_uploader(t("xlsx（可多选）"), type=["xlsx"],
                           accept_multiple_files=True, key="batch_up")
    _bm_ok = banma_client.is_configured()
    # 斑马 token の期限警告（Boss 2026-08-29 に定時同期を停止した結果、
    # token は「請求書を上げた時」しか更新されない。RefreshToken 30 日を
    # 跨いで空けると全失効 → ERP 画面で人手更新が必要になる）
    if _bm_ok:
        try:
            _tk = conn.execute(
                "SELECT refresh_expiry FROM banma.api_token "
                "WHERE app_id = %s", (banma_client._secret("BANMA_APP_ID"),)
            ).fetchone()
            if _tk and _tk["refresh_expiry"]:
                # 期限は中国標準時 naive で保存されている
                _now_cn = datetime.now(timezone(timedelta(hours=8))).replace(tzinfo=None)
                _left = (_tk["refresh_expiry"] - _now_cn).days
                _exp = _tk["refresh_expiry"].strftime("%Y-%m-%d")
                if _left < 0:
                    st.error(t("🔴 斑马 RefreshToken は {d} に失効済。斑马 ERP（服务 > 开放平台 > APP管理）で"
                               "「更新令牌」を押してから再度アップロードしてください。").format(d=_exp))
                elif _left <= 10:
                    st.warning(t("⚠️ 斑马 RefreshToken の残り {n} 日（{d} 失効）。"
                                 "この画面で請求書を上げるたびに自動延長されます。"
                                 "失効した場合は ERP 画面で「更新令牌」。").format(n=_left, d=_exp))
                else:
                    st.caption(t("🦓 斑马 token 有効（残り {n} 日 · {d} まで · アップロードのたび自動延長）")
                               .format(n=_left, d=_exp))
        except Exception:  # noqa: BLE001
            try:
                conn.rollback()
            except Exception:
                pass
    use_banma = st.checkbox(
        t("🦓 自动从斑马补齐 包裹→店铺（替代 BM 手动导出）"),
        value=_bm_ok, disabled=not _bm_ok, key="use_banma",
        help=t("按请求书费用发生日期 ±10 天从斑马拉包裹（全平台 · 约 10 分钟/月）。"
               "未配置 BANMA_APP_ID/SECRET 时不可用。"))
    if ups and st.button(t("💾 批量解析 → 写入 PG + 重算"), key="batch_btn", type="primary"):
        inv_log, bm_log, err_log = [], [], []
        inv_months: dict[str, list] = {}          # ym → 費用発生日 list（斑马窓計算用）
        prog = st.progress(0.0, text=t("解析中…"))
        for i, up in enumerate(ups):
            data = up.getvalue()
            try:
                kind, ym = detect_kind(data, up.name)
                if kind == "invoice" and not ym:
                    err_log.append(t("❌ {f}: 无法识别月份（文件名无 YYYYMM）").format(f=up.name))
                elif kind == "invoice":
                    rows, per = parse_invoice(data, ym)
                    if not rows:
                        err_log.append(t("❌ {f}: 录入0行").format(f=up.name))
                    else:
                        for ct in {r[1] for r in rows}:
                            conn.execute("DELETE FROM logistics.cost_invoice_raw WHERE year_month=%s AND cost_type=%s", (ym, ct))
                        conn.executemany(
                            """INSERT INTO logistics.cost_invoice_raw
                               (year_month, cost_type, join_key, amount_ex_tax, amount_in_tax,
                                material_cd, material_qty, sku, cost_date)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""", rows)
                        bill = parse_billing(data, ym)
                        if bill:
                            conn.execute("DELETE FROM logistics.cost_billing WHERE year_month=%s", (ym,))
                            conn.executemany(
                                """INSERT INTO logistics.cost_billing
                                   (year_month, seq, item_name, amount_ex_tax, amount_in_tax)
                                   VALUES (%s,%s,%s,%s,%s)""", bill)
                        conn.commit()
                        inv_months.setdefault(ym, []).extend(r[8] for r in rows)
                        inv_log.append(t("✅ 请求书 {ym}: ").format(ym=ym)
                                       + " ".join(f"{k}={v}" for k, v in per.items())
                                       + (f" +Billing{len(bill)}" if bill else ""))
                elif kind == "bm":
                    bm_rows = parse_bm(data)
                    if not bm_rows:
                        err_log.append(t("❌ {f}: 包裹0件").format(f=up.name))
                    else:
                        conn.executemany(
                            """INSERT INTO logistics.order_shop_map
                               (parcel_no, order_id, waybill_no, platform, shop, ship_date)
                               VALUES (%s,%s,%s,%s,%s,%s)
                               ON CONFLICT (parcel_no) DO UPDATE SET
                                 order_id=EXCLUDED.order_id, waybill_no=EXCLUDED.waybill_no,
                                 platform=EXCLUDED.platform, shop=EXCLUDED.shop,
                                 ship_date=EXCLUDED.ship_date, imported_at=now()""", bm_rows)
                        conn.commit()
                        bm_log.append(t("✅ BM {f}: {n}包裹 / {s}店铺").format(
                            f=up.name, n=len(bm_rows), s=len({r[4] for r in bm_rows if r[4]})))
                else:
                    err_log.append(t("❌ {f}: 无法判定类别（既非请求书也非 BM）").format(f=up.name))
            except Exception as e:
                err_log.append(f"❌ {up.name}: {e}")
            prog.progress((i + 1) / len(ups), text=f"{i + 1}/{len(ups)}")

        # ── 斑马 API 補齊（Boss 2026-08-29「先入库、再只拉需要的」）──
        #    請求書の join_key のうち未マッチ分だけを IDs/OrderDisplayID で
        #    精確批量取得（200 個/批）。窓方式（fill_shop_map_from_banma）は
        #    回灌用の後備として残置。
        if use_banma and inv_months:
            banma_client.ensure_store_map_table(conn)
            try:
                keys = banma_client.missing_join_keys(conn, sorted(inv_months))
                if keys:
                    bprog = st.progress(0.0, text=t("🦓 斑马 精确取数中…"))

                    def _cb(done, total):
                        bprog.progress(min(done / max(total, 1), 1.0),
                                       text=t("🦓 斑马: {p}/{n} 批").format(
                                           p=done, n=total))
                    try:
                        r = banma_client.fetch_shop_map_by_keys(conn, keys, _cb)
                        bm_log.append(t(
                            "🦓 斑马: 未匹配 {k} 单号 → 取得 {f} / 补齐 {u} 包裹（{b} 批）"
                        ).format(k=f"{r['requested']:,}", f=f"{r['fetched']:,}",
                                 u=f"{r['upserted']:,}", b=r["batches"]))
                    finally:
                        bprog.empty()
                else:
                    bm_log.append(t("🦓 斑马: 请求书单号已全部匹配，无需取数"))
            except banma_client.BanmaAuthError as e:
                err_log.append(t("❌ 斑马 token 失效（去 ERP 后台手动更新）: ") + str(e))
            except Exception as e:  # noqa: BLE001
                try:
                    conn.rollback()
                except Exception:
                    pass
                err_log.append(t("❌ 斑马补齐失败（已入库数据保留 · 重新点按钮即续跑）: ")
                               + str(e))

        # ── SO 形状（NST 注文）の自動帰類 ──
        #    Boss 2026-08-30 是正: shop には NST の店舗字段を使う（顧客名は
        #    販売渠道ではない）。店舗未設定の直録注文は「NST直販」（dept=EC 登録済）。
        if inv_months:
            try:
                so_keys = [k for k in banma_client.missing_join_keys(
                               conn, sorted(inv_months))
                           if re.match(r"^SO\d+", k)]
                if so_keys and nst_suiteql.is_configured():
                    names = nst_suiteql.lookup_so_shops(
                        [k.split("_")[0] for k in so_keys])
                    so_rows = [{"parcel_no": k, "order_id": k.split("_")[0],
                                "waybill_no": None, "platform": "NST",
                                "shop": names[k.split("_")[0]],
                                "ship_date": None}
                               for k in so_keys if k.split("_")[0] in names]
                    if so_rows:
                        conn.cursor().executemany(
                            banma_client.UPSERT_SHOP_MAP, so_rows)
                        conn.commit()
                        bm_log.append(t(
                            "🏢 NST 直録(SO): {n} 件 → 顧客名で登記（部署未分類なら tab② で分類）"
                        ).format(n=len(so_rows)))
                    left = [k for k in so_keys if k.split("_")[0] not in names]
                    if left:
                        err_log.append(t("⚠️ NST に見つからない SO: ") + ", ".join(left))
            except Exception as e:  # noqa: BLE001
                try:
                    conn.rollback()
                except Exception:
                    pass
                err_log.append(t("❌ SO 帰類失败（該当行は【不明】のまま）: ") + str(e))

        # ── 佐川 12 位運単（斑马体系外の国内出荷）→ 佐川直送/EC ──
        #    Boss 2026-08-30 拍板「佐川单都放在CB部门」。斑马/SO で拾えず
        #    残った 12 位純数字はここで自動帰類（shop_dept_map に EC 登録済）。
        if inv_months:
            try:
                sagawa = [k for k in banma_client.missing_join_keys(
                              conn, sorted(inv_months))
                          if re.fullmatch(r"\d{12}",
                                          banma_client.strip_seq_suffix(k))]
                if sagawa:
                    conn.cursor().executemany(
                        banma_client.UPSERT_SHOP_MAP,
                        [{"parcel_no": k, "order_id": None, "waybill_no": k,
                          "platform": "Sagawa", "shop": "佐川直送",
                          "ship_date": None} for k in sagawa])
                    conn.commit()
                    bm_log.append(t("🚛 佐川直送(EC): {n} 運単を帰類").format(
                        n=len(sagawa)))
            except Exception as e:  # noqa: BLE001
                try:
                    conn.rollback()
                except Exception:
                    pass
                err_log.append(t("❌ 佐川帰類失败: ") + str(e))

        with st.spinner(t("全月 重算中…")):
            run_recompute(None)
        if inv_log:
            st.success(t("请求书 {n} 件").format(n=len(inv_log)) + "\n\n" + "\n\n".join(inv_log))
        if bm_log:
            st.success(t("BM {n} 件").format(n=len(bm_log)) + "\n\n" + "\n\n".join(bm_log))
        if err_log:
            st.warning("\n\n".join(err_log))
        show_match_feedback()

# ============================================================
# tab2 · 店舗 → 部署 分類
# ============================================================
with tab2:
    st.markdown(t("##### 店铺 → 部署（输出 / EC）"))
    st.caption(t("未分类的店铺会计入【不明】。分类 → 保存 → 重算后反映到 输出/EC。"))

    unknown = pd.DataFrame([dict(r) for r in conn.execute(
        """SELECT DISTINCT m.shop AS shop
           FROM logistics.order_shop_map m
           LEFT JOIN logistics.shop_dept_map d ON m.shop = d.shop
           WHERE d.shop IS NULL AND m.shop IS NOT NULL
           ORDER BY m.shop"""
    ).fetchall()])

    if not unknown.empty:
        st.warning(t("⚠️ 未分类 {n} 店铺（当前计入【不明】）").format(n=len(unknown)))
        unknown["dept"] = "輸出"
        edited_new = st.data_editor(
            unknown, hide_index=True, key="new_dept",
            column_config={
                "shop": st.column_config.TextColumn(t("店铺"), disabled=True),
                "dept": st.column_config.SelectboxColumn(t("部署"), options=["輸出", "EC"]),
            },
        )
        if st.button(t("➕ 登记 + 重算"), key="new_dept_btn", type="primary"):
            conn.executemany(
                """INSERT INTO logistics.shop_dept_map (shop, dept) VALUES (%s,%s)
                   ON CONFLICT (shop) DO UPDATE SET dept=EXCLUDED.dept, updated_at=now()""",
                [(r["shop"], r["dept"]) for _, r in edited_new.iterrows()],
            )
            conn.commit()
            run_recompute(None)
            st.success(t("✅ 登记 + 重算完成"))
            st.rerun()
    else:
        st.success(t("无未分类店铺 ✅"))

    st.divider()
    st.markdown(t("##### 既有映射"))
    cur_map = pd.DataFrame([dict(r) for r in conn.execute(
        "SELECT shop, dept FROM logistics.shop_dept_map ORDER BY dept, shop"
    ).fetchall()])
    if cur_map.empty:
        st.info(t("还没有映射。请从上方未分类登记。"))
    else:
        edited = st.data_editor(
            cur_map, hide_index=True, num_rows="dynamic", key="edit_dept",
            column_config={
                "shop": st.column_config.TextColumn(t("店铺")),
                "dept": st.column_config.SelectboxColumn(t("部署"), options=["輸出", "EC"]),
            },
        )
        if st.button(t("💾 保存 + 重算"), key="edit_dept_btn"):
            conn.execute("DELETE FROM logistics.shop_dept_map")
            conn.executemany(
                """INSERT INTO logistics.shop_dept_map (shop, dept) VALUES (%s,%s)
                   ON CONFLICT (shop) DO UPDATE SET dept=EXCLUDED.dept, updated_at=now()""",
                [(r["shop"], r["dept"]) for _, r in edited.iterrows()
                 if r.get("shop") and r.get("dept")],
            )
            conn.commit()
            run_recompute(None)
            st.success(t("✅ 保存 + 重算完成"))
            st.rerun()

# ============================================================
# tab3 · Coupang 費用（LBF 輸送費明細 + 其他请求 手入力 · Boss 2026-07-14）
#   Coupang は後期 JDL 発送を使わない（JDL は梱包のみ → Pick&Pack/包材は JD 請求書側）。
#   輸送費は「YYYYMMLBF様輸送費.xlsx」で別途到着 → 暂时只读 K 列（Total JPY）合計。
#   【其他请求】= Boss 手動入力項目。合計 = 輸送費 + 其他请求。
# ============================================================
with tab3:
    st.markdown(t("##### Coupang 费用明细（LBF 輸送費 + 其他请求）"))
    st.caption(t(
        "上传「YYYYMM…LBF様輸送費.xlsx」→ 读 K 列（Total JPY）合计入库 · 月份取自文件名 YYYYMM · "
        "【其他请求】手动输入 · 合计 = 輸送費 + 其他请求 · 与 JD 配賦（包装/包材）相互独立"
    ))

    def _ensure_coupang_cost() -> str | None:
        try:
            conn.execute(
                "CREATE TABLE IF NOT EXISTS logistics.coupang_cost_monthly ("
                "year_month TEXT PRIMARY KEY, "
                "freight_amount NUMERIC(14,2), freight_rows INTEGER, "
                "other_amount NUMERIC(14,2), "
                "updated_at TIMESTAMPTZ DEFAULT NOW())")
            conn.commit()
            return None
        except Exception as e:  # noqa: BLE001
            try:
                conn.rollback()
            except Exception:
                pass
            return str(e)

    _cp_err = _ensure_coupang_cost()
    if _cp_err:
        st.error(t("⚠️ logistics.coupang_cost_monthly 初始化失败: ") + _cp_err)
        st.stop()

    # ── ① LBF 輸送費上传（K列合計）──
    cp_ups = st.file_uploader(t("LBF 輸送費 xlsx（可多选）"), type=["xlsx"],
                              accept_multiple_files=True, key="cp_up")
    if cp_ups and st.button(t("💾 解析 → 写入"), key="cp_btn", type="primary"):
        _oks, _errs = [], []
        for up in cp_ups:
            try:
                m = re.search(r"(20\d{2})\D?(\d{2})", up.name)
                if not m:
                    _errs.append(t("❌ {f}: 无法识别月份（文件名无 YYYYMM）").format(f=up.name))
                    continue
                _cym = f"{m[1]}-{m[2]}"
                # dimension 情報が壊れた出力あり → read_only 不可（1行しか読めない実績）
                wb = openpyxl.load_workbook(io.BytesIO(up.getvalue()), data_only=True)
                _tot, _n = 0.0, 0
                for ws in wb.worksheets:
                    for r in ws.iter_rows(min_col=11, max_col=11, values_only=True):
                        v = _num(r[0])
                        if v is not None:
                            _tot += v
                            _n += 1
                wb.close()
                if _n == 0:
                    _errs.append(t("❌ {f}: K列无数值").format(f=up.name))
                    continue
                conn.execute(
                    """INSERT INTO logistics.coupang_cost_monthly
                       (year_month, freight_amount, freight_rows, updated_at)
                       VALUES (%s,%s,%s,now())
                       ON CONFLICT (year_month) DO UPDATE SET
                         freight_amount=EXCLUDED.freight_amount,
                         freight_rows=EXCLUDED.freight_rows, updated_at=now()""",
                    (_cym, _tot, _n))
                conn.commit()
                _oks.append(t("✅ {ym}: 輸送費 {a}（{n} 件）").format(
                    ym=_cym, a=f"¥{_tot:,.0f}", n=f"{_n:,}"))
            except Exception as e:  # noqa: BLE001
                try:
                    conn.rollback()
                except Exception:
                    pass
                _errs.append(f"❌ {up.name}: {e}")
        if _oks:
            st.success("\n\n".join(_oks))
        if _errs:
            st.warning("\n\n".join(_errs))

    st.divider()

    # ── ②【其他请求】手入力（Boss 専用 · 月別）──
    st.markdown("##### " + t("✏️ 其他请求（手动输入）"))

    def _months_back(n: int = 14) -> list[str]:
        _t = date.today()
        y, mo = _t.year, _t.month
        out = []
        for _ in range(n):
            out.append(f"{y}-{mo:02d}")
            mo -= 1
            if mo == 0:
                y, mo = y - 1, 12
        return out

    _sel_ym = st.selectbox(t("対象月"), _months_back(), key="cp_other_ym")
    _cur = None
    try:
        _cur = conn.execute(
            "SELECT other_amount FROM logistics.coupang_cost_monthly "
            "WHERE year_month=%s", (_sel_ym,)).fetchone()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    _other_val = (float(_cur["other_amount"])
                  if _cur is not None and _cur["other_amount"] is not None else 0.0)
    _other = st.number_input(t("其他请求（円）"), min_value=0.0, value=_other_val,
                             step=1000.0, format="%.0f", key=f"cp_other_{_sel_ym}")
    if st.button(t("💾 保存"), key=f"cp_other_btn_{_sel_ym}"):
        conn.execute(
            """INSERT INTO logistics.coupang_cost_monthly
               (year_month, other_amount, updated_at)
               VALUES (%s,%s,now())
               ON CONFLICT (year_month) DO UPDATE SET
                 other_amount=EXCLUDED.other_amount, updated_at=now()""",
            (_sel_ym, _other))
        conn.commit()
        st.rerun()

    # ── ③ 月別一覧（輸送費 → 其他请求 → 合計）──
    _all = conn.execute(
        "SELECT year_month, freight_amount, freight_rows, other_amount "
        "FROM logistics.coupang_cost_monthly ORDER BY year_month DESC").fetchall()
    if _all:
        _cdf = pd.DataFrame([dict(r) for r in _all])
        for _c in ("freight_amount", "other_amount"):
            _cdf[_c] = pd.to_numeric(_cdf[_c], errors="coerce")
        _cdf["total"] = _cdf[["freight_amount", "other_amount"]].fillna(0).sum(axis=1)

        def _y(v):
            return "—" if pd.isna(v) else f"¥{v:,.0f}"

        st.dataframe(pd.DataFrame({
            t("月"): _cdf["year_month"],
            t("輸送費(K列合计)"): _cdf["freight_amount"].map(_y),
            t("件数"): _cdf["freight_rows"].map(
                lambda v: "—" if pd.isna(v) else f"{int(v):,}"),
            t("其他请求"): _cdf["other_amount"].map(_y),
            t("合计"): _cdf["total"].map(_y),
        }), hide_index=True, use_container_width=True)
