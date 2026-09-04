"""運営の実ファイル 2 本を突き合わせる回帰テスト。

fixtures は 2026-09-02 に運営が実際に流した 37 件:
    coupang_orders_20260902.json  … Coupang からダウンロードした受注（入力）
    ecms_upload_20260902.json     … それを整形して ECMS に上げたもの（正解）

**この 2 本が正**。変換規則をいじって落ちたら、いじった方が間違っている。
氏名・電話・住所・PCCC が入っているので fixtures は社外に出さないこと。
"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import pytest

from shared import coupang_to_ecms_xlsx as X

FIX = Path(__file__).parent / "fixtures"
SRC = json.loads((FIX / "coupang_orders_20260902.json").read_text("utf-8"))[1:]
DST = json.loads((FIX / "ecms_upload_20260902.json").read_text("utf-8"))[1:]
BY_ORDER = {d["B"]: d for d in DST}


def _products() -> dict:
    """正解ファイルから商品マスタを逆算（英語品名 / HScode / ProductID / OptionID）。

    本番では coupang_product_info テーブルから引く。ここは変換規則の検証が目的なので、
    マスタが正しく引けた前提を置いて**それ以外の列**を突き合わせる。
    """
    out = {}
    for s_row in SRC:
        d = BY_ORDER.get(s_row.get("C"))
        if not d:
            continue
        pid = (d.get("AN", "").split("/products/")[-1].split("?")[0]
               if "/products/" in d.get("AN", "") else "")
        # キーは SKU。同じ JAN でも規格違いは別 OptionID・別英語品名になる
        out[(s_row.get("Q") or "").strip()] = {
            "name_en": d.get("AH", ""), "hscode": d.get("AM", ""),
            "product_id": pid, "option_id": d.get("AU", "")}
    return out


def _masters() -> dict:
    """NST 商品マスタの**実値**（`cms0811` シートから抜いた 37 JAN 分の 厂商 / 毛重(g)）。

    正解ファイルから逆算すると自己参照になって検証にならないので、別ソースの実値を使う。
    """
    return json.loads((FIX / "nst_master_20260902.json").read_text("utf-8"))


PRODUCTS, MASTERS = _products(), _masters()


def _converted() -> list[dict]:
    return X.convert(SRC, PRODUCTS, MASTERS, start_seq=1, on=date(2026, 9, 2))


def test_行数が一致():
    assert len(SRC) == 37 and len(DST) == 37
    assert len(_converted()) == 37


def test_毛重は_NST_の_g_を_kg_に丸めたもの():
    """AJ = round(毛重(g) ÷ 1000, 2)。**個数も購入数も掛けない**。

    実測: 148g→0.15 / 83g→0.08 / 85g→0.09 / 757g→0.76。
    ここを掛け算にすると申告重量が数倍になり、ECMS 側の実重量と合わなくなる。
    """
    bad = []
    for got in _converted():
        want = BY_ORDER[got["B"]]
        if str(got["AJ"]) != str(want["AJ"]):
            bad.append((got["AG"], got["AJ"], want["AJ"]))
    assert not bad, f"毛重 不一致 {len(bad)}/37: {bad[:5]}"


def test_品牌は_厂商から和名併記を落とす():
    """`Pelican Soap（ペリカン石鹸）` → `Pelican Soap`。これで 32/37。

    残り 5 行（4 種）は NST のメーカー名と運営のブランド名が別物で、自動では導けない:
        ユニリーバ → Dove / コーセーコスメポート → CoenRich
        UHA味覚糖 → UHA / マルコメ → Marukome
    これらは商品マスタの `brand` 列で上書きする運用。**推測で寄せない**。
    """
    hit = sum(1 for g in _converted() if g["AE"] == BY_ORDER[g["B"]]["AE"])
    assert hit == 32, f"品牌 {hit}/37（想定 32）"
    assert X.clean_brand("Pelican Soap（ペリカン石鹸）") == "Pelican Soap"
    assert X.clean_brand("IDA Laboratories（井田ラボラトリー）") == "IDA Laboratories"
    assert X.clean_brand("KAO") == "KAO"


def test_HSCode_は載せない():
    """運営の実ファイルは 37/37 とも空。マスタに値があっても出さない。"""
    row = X.build_row({X.C_SKU: "4901616011007_3"}, {"hscode": "330610"}, {}, "R1")
    assert row["AM"] == ""


def test_옵션ID_は注文側が正():
    """商品マスタが古いことがある。実測 37/37 が注文の Option ID と一致。"""
    row = X.build_row({X.C_SKU: "x_1", X.C_OPTION_ID: "95323351399"},
                      {"option_id": "95424664036", "product_id": "9544746544"}, {}, "R1")
    assert row["AU"] == "95323351399"
    assert "vendorItemId=95323351399" in row["AN"]
    assert "/products/9544746544" in row["AN"]      # Product ID はマスタ側


def test_商品マスタの_brand_が_厂商より優先():
    row = X.build_row({X.C_SKU: "4902111775227_1"},
                      {"brand": "Dove", "name_en": "x"}, {"maker": "ユニリーバ"}, "R1")
    assert row["AE"] == "Dove"


def test_Excelと同じ四捨五入():
    """組み込み round は偶数丸めで 0.015 → 0.01。運営の実出力は 0.02（NST 15g）。"""
    assert X.round_half_up(0.015) == 0.02
    assert X.round_half_up(0.085) == 0.09
    assert X.round_half_up(0.0849) == 0.08


@pytest.mark.parametrize("col,label", [
    ("A", "Client Code"), ("B", "订单号"), ("L", "仓库编码"), ("Q", "收货人语种"),
    ("R", "姓名"), ("S", "电话"), ("U", "国家"), ("Y", "地址"), ("Z", "证件类型"),
    ("AA", "PCCC"), ("AC", "内件语种"), ("AF", "规格型号"),
    ("AG", "SKU"), ("AH", "英文品名"), ("AK", "重量单位"), ("AL", "危险品"),
    ("AN", "商品URL"), ("AO", "数量"), ("AR", "币种"), ("AS", "原产国"),
    ("AT", "发货人编码"), ("AU", "Platform Id"),
])
def test_列が実出力と一致(col, label):
    """37 行すべてで運営の出力と同じ値になること。"""
    bad = []
    for got in _converted():
        want = BY_ORDER[got["B"]]
        if str(got.get(col, "")) != str(want.get(col, "")):
            bad.append((got["B"], got.get(col), want.get(col)))
    assert not bad, f"{label}({col}) 不一致 {len(bad)}/37: {bad[:3]}"


def test_単価は韓国ウォンの整数():
    """USD 換算しない（× 0.00068 は JD 向けの別ライン）。かつ**整数に四捨五入**。

    実測: 34720÷3 = 11573.33 → 11573 / 20600÷3 = 6866.67 → **6867**。
    切り捨てで書くと 1 ウォンずつズレて申告額が合わなくなる。
    """
    for got in _converted():
        want = BY_ORDER[got["B"]]
        assert got["AR"] == "KRW"
        assert isinstance(got["AP"], int), got["B"]
        assert got["AP"] == int(want["AP"]), (got["B"], got["AP"], want["AP"])


def test_邮编は数値():
    """運営の実ファイルは 0902・0903 とも**数値**（`05564` ではなく `5564`）。

    前ゼロは落ちるが ECMS はそれで通っている。文字列で出すと運営の出力と食い違う。
    """
    for g in _converted():
        assert isinstance(g["X"], int), (g["B"], g["X"])
        assert g["X"] == int(BY_ORDER[g["B"]]["X"])


def test_省と市_36件一致():
    """残る 1 件は `전남광주통합특별시` という実在しない行政区。人が直したもので、
    自動では埋めない（埋めると誤った通関先になる）。"""
    v_ok = w_ok = 0
    diffs = []
    for got in _converted():
        want = BY_ORDER[got["B"]]
        v_ok += got["V"] == want["V"]
        w_ok += got["W"] == want["W"]
        if got["V"] != want["V"] or got["W"] != want["W"]:
            diffs.append((got["Y"][:24], got["V"], want["V"], got["W"], want["W"]))
    assert w_ok == 36, f"市 {w_ok}/37 · {diffs}"
    assert v_ok == 36, f"省 {v_ok}/37 · {diffs}"


def test_改称前の道名は書き換えない():
    """`전라북도` を `전북특별자치도` に直さない——運営はそのまま出している。"""
    assert X.province_city("전라북도 전주시 완산구 평화동2가 896-1")[:2] == ("전라북도", "전주시")


def test_略称は正式名に開く():
    assert X.province_city("서울 강서구 가양동 1487")[:2] == ("서울특별시", "강서구")


def test_市に区を混ぜない():
    """`수원시 영통구` は市＝수원시。구を入れると ECMS 側の市名と合わない。"""
    assert X.province_city("경기도 수원시 영통구 도청로 65")[:2] == ("경기도", "수원시")


def test_未知の行政区は空にする():
    p, c, how = X.province_city("전남광주통합특별시 북구 자미로39번길 9")
    assert c == "", "実在しない行政区から市を作ってはいけない"
    assert how == "no_sgg"


def test_頭程運単号の形():
    assert X.ref_number(1, date(2026, 9, 2)) == "ECLBF26090200001"
    assert X.ref_number(40, date(2026, 9, 2)) == "ECLBF26090200040"
    refs = [g["C"] for g in _converted()]
    assert len(set(refs)) == 37, "連番が重複している"


def test_必須欠けを拾う():
    rows = _converted()
    assert X.missing(rows[0]) == []
    holed = dict(rows[0], AA="", AH="")
    assert set(X.missing(holed)) == {"PCCC", "英文品名"}


def test_xlsx_書き出し(tmp_path):
    p = X.to_xlsx(_converted(), tmp_path / "out.xlsx")
    from openpyxl import load_workbook
    ws = load_workbook(p).active
    assert ws.max_column == 57
    assert ws.max_row == 38                                   # ヘッダ + 37
    assert ws.cell(1, 1).value.startswith("Client Code")
    assert ws.cell(2, 1).value == "LBF"
    zip_col = X.COLUMNS.index("X") + 1
    assert isinstance(ws.cell(2, zip_col).value, int)         # 郵便番号は数値
    sku_col = X.COLUMNS.index("AG") + 1
    assert isinstance(ws.cell(2, sku_col).value, str)         # SKU は文字列


def test_品牌の日英対応():
    """運営「品牌（要填英文的）」。メーカー単位で 1 行入れれば全商品に効く。"""
    assert X.needs_english_brand("コーセーコスメポート") is True
    assert X.needs_english_brand("UHA味覚糖") is True
    assert X.needs_english_brand("Pelican Soap") is False
    assert X.needs_english_brand("") is False
    row = X.build_row({X.C_SKU: "x_1"}, {}, {"maker": "ユニリーバ"}, "R",
                      {"ユニリーバ": "Dove"})
    assert row["AE"] == "Dove"
    # 商品マスタの brand が最優先
    row2 = X.build_row({X.C_SKU: "x_1"}, {"brand": "Unilever"}, {"maker": "ユニリーバ"}, "R",
                       {"ユニリーバ": "Dove"})
    assert row2["AE"] == "Unilever"


def test_毛重はマスタの入数込み重量を割る():
    """「产品重量」は SKU 全体（入数込み）の kg。ECMS の Item_Grossweight は 1 個あたり。

    実測: 4901616011007_3 の 0.444kg ÷ 3 = 0.148 → 0.15（運営の実出力と一致）。
    割らずに出すと 3 倍の重量で申告することになる。
    """
    row = X.build_row({X.C_SKU: "4901616011007_3"}, {"weight_kg": 0.444}, {}, "R")
    assert row["AJ"] == 0.15
    # マスタに無ければ NST（g）へフォールバック
    assert X.build_row({X.C_SKU: "x_2"}, {}, {"weight": 148.0}, "R")["AJ"] == 0.15
    # どちらも無ければ空。0 で埋めない
    assert X.build_row({X.C_SKU: "x_1"}, {}, {}, "R")["AJ"] == ""


def test_世宗市と新設区():
    """世宗は単層制で시군구が無い。검단구は 2026-07 新設で手元の법정동코드に載っていない。"""
    assert X.province_city("세종특별자치시 마음로 67 가락마을")[:2] == \
        ("세종특별자치시", "세종시")      # 市欄は略称（運営の実出力）
    assert X.province_city("인천광역시 검단구 원당동 987-1")[:2] == ("인천광역시", "검단구")


def test_実在しない行政区は省も空にする():
    """`전남광주통합특별시` は前方一致で전라남도に化ける。実際は광주광역시で通関先が変わる。"""
    p, c, _ = X.province_city("전남광주통합특별시 북구 자미로39번길 9")
    assert (p, c) == ("", "")


# ------------------------------------------------------------------
# 0903（2 組目の正解）· 0902 だけでは出なかった形が入っている
# ------------------------------------------------------------------
SRC03 = json.loads((FIX / "coupang_orders_20260903.json").read_text("utf-8"))[1:]
DST03 = json.loads((FIX / "ecms_upload_20260903.json").read_text("utf-8"))[1:]
PROD03 = json.loads((FIX / "coupang_products_20260903.json").read_text("utf-8"))
BY03 = {d["B"]: d for d in DST03}


def _converted03():
    from datetime import date as _d
    return X.convert(SRC03, PROD03, {}, start_seq=1, on=_d(2026, 9, 3))


@pytest.mark.parametrize("col,label", [
    ("B", "订单号"), ("R", "姓名"), ("S", "电话"), ("X", "邮编"), ("Y", "地址"),
    ("AA", "PCCC"), ("V", "省"), ("W", "市"), ("AF", "规格型号"), ("AG", "SKU"),
    ("AJ", "毛重"), ("AO", "数量"), ("AP", "单价"), ("AU", "Platform Id"),
])
def test_0903も実出力と一致(col, label):
    bad = []
    for got in _converted03():
        want = BY03[got["B"]]
        if str(got.get(col, "")) != str(want.get(col, "")):
            bad.append((got["B"], got.get(col), want.get(col)))
    assert not bad, f"{label}({col}) 不一致 {len(bad)}/{len(DST03)}: {bad[:3]}"


def test_毛重は内件総数で割る():
    """0903 で初めて出た形: 入数 1 × 購入 2。0902 は購入数が全部 1 で区別が付かなかった。

    4902111773421（入数 1 · 購入 2 · 内件総数 2）: 0.23 / 2 = 0.115 → 0.12
    入数で割ると 0.23 のままになり、運営の出力と食い違う。
    """
    row = X.build_row({X.C_SKU: "4902111773421", X.C_QTY: "2"},
                      {"weight_kg": 0.23}, {}, "R")
    assert row["AO"] == 2 and row["AJ"] == 0.12


def test_世宗の市欄は略称():
    """運営の実出力は `세종시`。`세종특별자치시` をそのまま入れない。"""
    assert X.province_city("세종특별자치시 마음로 67 가락마을")[:2] == ("세종특별자치시", "세종시")


# ------------------------------------------------------------------
# PCCC 形式チェック（2026-09-04 許慧杰さん要望）
# ------------------------------------------------------------------
def test_PCCC_形式NGは送る前に落とす():
    """許慧杰さん運用『通关号码不对的话上传ecms系统之前会摘掉』と同じ挙動。

    ⚠️ これは P+12桁の**形式**だけを見る。氏名・電話（・2026-02 以降は郵便番号）が
    本人と一致するかは 韓国関税庁 UNIPASS の Open API でしか検証できず未接続——
    「桁が合っているだけの偽物」は弾けるが「別人の正しい番号の誤入力」は弾けない。
    """
    for bad in ("ABC12345", "P84216010747", "P8421601074766", "", "  ", "P" + "1" * 11 + "A"):
        row = X.build_row({X.C_SKU: "x_1", X.C_PCCC: bad}, {}, {}, "R")
        assert row["AA"] == "", f"{bad!r} は形式NGのはずなのに通した"


def test_PCCC_正常形式はそのまま():
    row = X.build_row({X.C_SKU: "x_1", X.C_PCCC: "P842160107476"}, {}, {}, "R")
    assert row["AA"] == "P842160107476"
    row2 = X.build_row({X.C_SKU: "x_1", X.C_PCCC: "p842160107476"}, {}, {}, "R")  # 小文字
    assert row2["AA"] == "p842160107476"       # 大小文字はそのまま出す（判定だけ緩める）


def test_pccc_dropped_で画面に理由を出せる():
    assert X.pccc_dropped({X.C_PCCC: "ABC12345"}) is True
    assert X.pccc_dropped({X.C_PCCC: "P842160107476"}) is False
    assert X.pccc_dropped({X.C_PCCC: ""}) is False       # 元から空はケース違い（missing 側で拾う）
